from io import BytesIO
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font,PatternFill

def excel_response(*,institucion,titulo,encabezados,filas,nombre,ciclo=None,filtros=""):
 wb=Workbook();ws=wb.active;ws.title=titulo[:31];ws.append([institucion.nombre]);ws.append([titulo]);ws.append([f"Generado: {timezone.localtime():%d/%m/%Y %H:%M}"]);ws.append([f"Ciclo: {ciclo}" if ciclo else "Todos los ciclos"]);ws.append([f"Filtros: {filtros or 'Sin filtros adicionales'}"]);ws.append([]);ws.append(list(encabezados))
 for cell in ws[7]:cell.font=Font(bold=True,color="FFFFFF");cell.fill=PatternFill("solid",fgColor="2563EB")
 for row in filas:ws.append(list(row))
 ws.freeze_panes="A8";ws.auto_filter.ref=ws.dimensions
 for col in ws.columns:ws.column_dimensions[col[0].column_letter].width=min(max(len(str(c.value or "")) for c in col)+2,45)
 out=BytesIO();wb.save(out);response=HttpResponse(out.getvalue(),content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet");response["Content-Disposition"]=f'attachment; filename="{nombre}"';return response
