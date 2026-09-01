from io import BytesIO
from django.core.exceptions import ValidationError
from django.db import transaction
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from academico.models import GradoInstitucion, Seccion
from .models import Alumno, AlumnoEncargado, Encargado, Familia, ImportacionAlumnos, Inscripcion

HEADERS=["CUI","PRIMER_NOMBRE","SEGUNDO_NOMBRE","OTROS_NOMBRES","PRIMER_APELLIDO","SEGUNDO_APELLIDO","FECHA_NACIMIENTO","SEXO","TELEFONO","EMAIL","DIRECCION","CODIGO_FAMILIA","CUI_ENCARGADO","NOMBRES_ENCARGADO","APELLIDOS_ENCARGADO","PARENTESCO","TELEFONO_ENCARGADO","EMAIL_ENCARGADO","PRINCIPAL","RESPONSABLE_FINANCIERO","CODIGO_OFERTA","CODIGO_GRADO","CODIGO_SECCION"]

def crear_plantilla(institucion,ciclo):
    wb=Workbook(); ws=wb.active; ws.title="ESTUDIANTES"; ws.append(HEADERS); ws.freeze_panes="A2"
    grados=wb.create_sheet("CATALOGO_GRADOS"); grados.append(["CODIGO_OFERTA","OFERTA","CODIGO_GRADO","GRADO"])
    for g in GradoInstitucion.objects.filter(institucion=institucion,ciclo=ciclo,activo=True).select_related("oferta"): grados.append([g.oferta.codigo_interno,g.oferta.nombre_mostrado,g.codigo,g.nombre])
    sec=wb.create_sheet("CATALOGO_SECCIONES"); sec.append(["CODIGO_GRADO","CODIGO_SECCION","SECCION","JORNADA"])
    for s in Seccion.objects.filter(institucion=institucion,ciclo=ciclo,activa=True).select_related("grado","jornada"): sec.append([s.grado.codigo,s.codigo,s.nombre,s.jornada.nombre if s.jornada else ""])
    ins=wb.create_sheet("INSTRUCCIONES"); ins.append(["PLANTILLA AULAPRO"]); ins.append(["Requeridos: primer nombre, primer apellido, fecha de nacimiento, sexo, oferta, grado y sección."]); ins.append(["CUI: 13 dígitos o vacío si está pendiente. Fecha: AAAA-MM-DD. Sexo: F, M u O."]); ins.append(["Use los códigos exactos de las hojas auxiliares. No agregue una columna de institución."])
    out=BytesIO(); wb.save(out); out.seek(0); return out

def prevalidar(archivo,institucion,ciclo):
    archivo.seek(0)
    try: wb=load_workbook(archivo,read_only=True,data_only=True)
    except Exception as exc: return {"filas":[],"errores":[f"Archivo XLSX inválido: {exc}"],"advertencias":[]}
    if "ESTUDIANTES" not in wb.sheetnames: return {"filas":[],"errores":["No existe la hoja ESTUDIANTES."],"advertencias":[]}
    ws=wb["ESTUDIANTES"]; headers=[str(v or "").strip().upper() for v in next(ws.iter_rows(values_only=True))]
    idx={h:i for i,h in enumerate(headers)}; faltantes=[h for h in HEADERS if h not in idx]
    if faltantes: return {"filas":[],"errores":["Faltan columnas: "+", ".join(faltantes)],"advertencias":[]}
    existentes={a.cui:a for a in Alumno.objects.filter(institucion=institucion,cui__isnull=False)}
    familias={f.codigo:f for f in Familia.objects.filter(institucion=institucion,codigo__isnull=False)}
    encargados={e.cui:e for e in Encargado.objects.filter(institucion=institucion,cui__isnull=False)}
    grados={(g.oferta.codigo_interno,g.codigo):g for g in GradoInstitucion.objects.filter(institucion=institucion,ciclo=ciclo).select_related("oferta")}
    secciones={(s.grado_id,s.codigo):s for s in Seccion.objects.filter(institucion=institucion,ciclo=ciclo).select_related("grado")}
    vistos=set(); filas=[]; errores=[]; advertencias=[]
    for numero,raw in enumerate(ws.iter_rows(min_row=2,values_only=True),2):
        if not any(v not in (None,"") for v in raw): continue
        data={h:raw[i] for h,i in idx.items()}; row_errors=[]; row_warn=[]
        cui=str(data["CUI"] or "").strip().replace(".0","")
        if cui and (not cui.isdigit() or len(cui)!=13): row_errors.append("CUI inválido")
        if cui and cui in vistos: row_errors.append("CUI repetido en el archivo")
        if cui: vistos.add(cui)
        for req in ("PRIMER_NOMBRE","PRIMER_APELLIDO","FECHA_NACIMIENTO","SEXO","CODIGO_OFERTA","CODIGO_GRADO","CODIGO_SECCION"):
            if not data[req]: row_errors.append(f"{req} es requerido")
        grado=grados.get((str(data["CODIGO_OFERTA"] or "").strip(),str(data["CODIGO_GRADO"] or "").strip()))
        if not grado: row_errors.append("Grado no encontrado")
        seccion=secciones.get((grado.pk,str(data["CODIGO_SECCION"] or "").strip())) if grado else None
        if grado and not seccion: row_errors.append("Sección no encontrada")
        alumno=existentes.get(cui) if cui else None
        if alumno and alumno.nombre_completo.lower()!=f'{data["PRIMER_NOMBRE"]} {data["PRIMER_APELLIDO"]}'.strip().lower(): row_warn.append("El CUI existe y los nombres no coinciden; no se actualizarán datos personales")
        if alumno and Inscripcion.objects.filter(alumno=alumno,ciclo=ciclo,estado=Inscripcion.Estado.ACTIVA).exists(): row_errors.append("El alumno ya tiene inscripción activa en este ciclo")
        fila={"numero":numero,"data":data,"cui":cui,"alumno":alumno,"grado":grado,"seccion":seccion,"familia":familias.get(str(data["CODIGO_FAMILIA"] or "").strip()),"encargado":encargados.get(str(data["CUI_ENCARGADO"] or "").strip()),"errores":row_errors,"advertencias":row_warn,"accion":"Alumno existente / nueva inscripción" if alumno else "Nuevo alumno"}
        filas.append(fila); errores.extend([f"Fila {numero}: {e}" for e in row_errors]); advertencias.extend([f"Fila {numero}: {e}" for e in row_warn])
    if not errores:
        from suscripciones.services import validar_cupo_alumnos
        try: validar_cupo_alumnos(institucion,len(filas))
        except ValidationError as exc: errores.extend(exc.messages)
    return {"filas":filas,"errores":errores,"advertencias":advertencias}

@transaction.atomic
def ejecutar_importacion(registro):
    resultado=prevalidar(registro.archivo_original,registro.institucion,registro.ciclo)
    if resultado["errores"]:
        registro.estado=ImportacionAlumnos.Estado.FALLIDA; registro.errores=len(resultado["errores"]); registro.detalle_errores=resultado["errores"]; registro.fecha_fin=timezone.now(); registro.save(); raise ValidationError("La importación contiene errores críticos.")
    creados=existentes=inscripciones=0
    for fila in resultado["filas"]:
        d=fila["data"]; alumno=fila["alumno"]
        if not alumno:
            alumno=Alumno.objects.create(institucion=registro.institucion,familia=fila["familia"],cui=fila["cui"] or None,primer_nombre=str(d["PRIMER_NOMBRE"]),segundo_nombre=str(d["SEGUNDO_NOMBRE"] or ""),otros_nombres=str(d["OTROS_NOMBRES"] or ""),primer_apellido=str(d["PRIMER_APELLIDO"]),segundo_apellido=str(d["SEGUNDO_APELLIDO"] or ""),fecha_nacimiento=d["FECHA_NACIMIENTO"],sexo=str(d["SEXO"]).upper(),telefono=str(d["TELEFONO"] or ""),email=str(d["EMAIL"] or ""),direccion=str(d["DIRECCION"] or ""),fecha_ingreso=timezone.localdate()); creados+=1
        else: existentes+=1
        enc_cui=str(d["CUI_ENCARGADO"] or "").strip() or None; encargado=fila["encargado"]
        if not encargado and d["NOMBRES_ENCARGADO"]:
            encargado=Encargado.objects.create(institucion=registro.institucion,cui=enc_cui,nombres=str(d["NOMBRES_ENCARGADO"]),apellidos=str(d["APELLIDOS_ENCARGADO"] or ""),telefono=str(d["TELEFONO_ENCARGADO"] or ""),email=str(d["EMAIL_ENCARGADO"] or ""))
        if encargado: AlumnoEncargado.objects.get_or_create(institucion=registro.institucion,alumno=alumno,encargado=encargado,defaults={"parentesco":str(d["PARENTESCO"] or "OTRO").upper(),"parentesco_otro":"No especificado" if str(d["PARENTESCO"] or "OTRO").upper()=="OTRO" else "","es_principal":str(d["PRINCIPAL"] or "").upper() in ("SI","SÍ","1","TRUE"),"es_responsable_financiero":str(d["RESPONSABLE_FINANCIERO"] or "").upper() in ("SI","SÍ","1","TRUE")})
        g=fila["grado"]; Inscripcion.objects.create(institucion=registro.institucion,alumno=alumno,ciclo=registro.ciclo,oferta_academica=g.oferta,grado=g,seccion=fila["seccion"],fecha_inscripcion=timezone.localdate()); inscripciones+=1
    registro.estado=ImportacionAlumnos.Estado.PROCESADA; registro.total_filas=len(resultado["filas"]); registro.creados=creados; registro.actualizados=existentes; registro.inscripciones_creadas=inscripciones; registro.errores=0; registro.fecha_fin=timezone.now(); registro.save(); return registro
from django.core.exceptions import ValidationError
from django.db import transaction
from django.utils import timezone

from academico.models import ResultadoAnualAlumno


@transaction.atomic
def reinscribir_alumno(*, resultado, ciclo_destino, seccion_destino):
    """Crea la inscripción anual de forma tenant-safe e idempotente."""
    from suscripciones.services import suscripcion_actual
    from .models import Inscripcion

    resultado = ResultadoAnualAlumno.objects.select_for_update().select_related("inscripcion", "alumno").get(pk=resultado.pk)
    if resultado.institucion_id != ciclo_destino.institucion_id or seccion_destino.institucion_id != resultado.institucion_id:
        raise ValidationError("Todos los datos deben pertenecer a la misma institución.")
    if seccion_destino.ciclo_id != ciclo_destino.pk:
        raise ValidationError("La sección no pertenece al ciclo destino.")
    permitidos = (ResultadoAnualAlumno.Resultado.PROMOVIDO, ResultadoAnualAlumno.Resultado.NO_PROMOVIDO)
    if resultado.resultado_final not in permitidos:
        raise ValidationError("El resultado confirmado no es elegible para reinscripción.")
    esperado = resultado.inscripcion.grado.orden + (1 if resultado.resultado_final == ResultadoAnualAlumno.Resultado.PROMOVIDO else 0)
    if seccion_destino.grado.orden != esperado:
        raise ValidationError("El grado destino no corresponde a la promoción propuesta.")
    existente = Inscripcion.objects.filter(alumno=resultado.alumno, ciclo=ciclo_destino, estado=Inscripcion.Estado.ACTIVA).first()
    if existente:
        return existente, False
    ocupados = Inscripcion.objects.select_for_update().filter(seccion=seccion_destino, estado=Inscripcion.Estado.ACTIVA).count()
    if seccion_destino.capacidad is not None and ocupados >= seccion_destino.capacidad:
        raise ValidationError("La sección destino no tiene capacidad disponible.")
    suscripcion = suscripcion_actual(resultado.institucion)
    usados = Inscripcion.objects.filter(institucion=resultado.institucion, ciclo=ciclo_destino, estado=Inscripcion.Estado.ACTIVA).count()
    if suscripcion and suscripcion.limite_alumnos is not None and usados + 1 > suscripcion.limite_alumnos:
        raise ValidationError("La reinscripción excede el límite de alumnos del plan.")
    inscripcion = Inscripcion.objects.create(institucion=resultado.institucion, alumno=resultado.alumno, ciclo=ciclo_destino,
        oferta_academica=seccion_destino.grado.oferta, grado=seccion_destino.grado, seccion=seccion_destino,
        fecha_inscripcion=timezone.localdate(), estado=Inscripcion.Estado.ACTIVA, es_reingreso=True)
    return inscripcion, True


@transaction.atomic
def reinscripcion_masiva(*, asignaciones):
    """Procesa (resultado, ciclo, sección) en una única transacción."""
    return [reinscribir_alumno(resultado=r, ciclo_destino=c, seccion_destino=s) for r, c, s in asignaciones]


def requisitos_aplicables(alumno, inscripcion=None, visible_portal=False):
    from django.db.models import Q
    from .models import RequisitoDocumentoAlumno
    inscripcion = inscripcion or alumno.inscripciones.filter(estado="ACTIVA").select_related("ciclo","oferta_academica__nivel","grado").first()
    qs=RequisitoDocumentoAlumno.objects.filter(institucion=alumno.institucion,activo=True,tipo_documento__activo=True).select_related("tipo_documento")
    if visible_portal:qs=qs.filter(tipo_documento__visible_portal=True)
    if not inscripcion:return qs.filter(aplica_a_nivel__isnull=True,aplica_a_oferta__isnull=True,aplica_a_grado__isnull=True,aplica_a_ciclo__isnull=True)
    return qs.filter(Q(aplica_a_nivel__isnull=True)|Q(aplica_a_nivel=inscripcion.oferta_academica.nivel),Q(aplica_a_oferta__isnull=True)|Q(aplica_a_oferta=inscripcion.oferta_academica),Q(aplica_a_grado__isnull=True)|Q(aplica_a_grado=inscripcion.grado),Q(aplica_a_ciclo__isnull=True)|Q(aplica_a_ciclo=inscripcion.ciclo)).distinct()


def resumen_expediente(alumno, inscripcion=None, visible_portal=False):
    from .models import DocumentoAlumno
    requisitos=list(requisitos_aplicables(alumno,inscripcion,visible_portal));items=[];aprobados=denominador=pendientes=rechazados=0
    for req in requisitos:
        documentos=alumno.documentos.filter(tipo_documento=req.tipo_documento).order_by("-fecha_carga")
        if req.aplica_a_ciclo_id:documentos=documentos.filter(ciclo_id=req.aplica_a_ciclo_id)
        documento=documentos.first();estado=documento.estado_vigente if documento else DocumentoAlumno.Estado.PENDIENTE
        obligatorio=req.obligatorio
        if estado==DocumentoAlumno.Estado.NO_APLICA:obligatorio=False
        if obligatorio:
            denominador+=1
            if estado==DocumentoAlumno.Estado.APROBADO:aprobados+=1
            else:pendientes+=1
        if estado==DocumentoAlumno.Estado.RECHAZADO:rechazados+=1
        items.append({"requisito":req,"documento":documento,"estado":estado,"obligatorio":obligatorio})
    porcentaje=round(aprobados*100/denominador) if denominador else 100
    return {"items":items,"aprobados":aprobados,"total":denominador,"pendientes":pendientes,"rechazados":rechazados,"porcentaje":porcentaje,"completo":porcentaje==100}


def documentos_por_vencer(institucion,dias=30):
    from datetime import timedelta
    from django.utils import timezone
    from .models import DocumentoAlumno
    hoy=timezone.localdate()
    return DocumentoAlumno.objects.filter(institucion=institucion,estado=DocumentoAlumno.Estado.APROBADO,fecha_vencimiento__range=(hoy,hoy+timedelta(days=dias)))
